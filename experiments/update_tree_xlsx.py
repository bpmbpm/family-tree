#!/usr/bin/env python3
"""
Script to update tree.xlsx with:
1. New person entries for siblings of Ilya Ulyanov and siblings of Maria Blank
2. New family entries for Ульянин+Смирнова and ensure Blank+Гроссшопф exists
3. New columns: birthPlace_, deathPlace_, burialPlace_, occupation_ for person sheet
4. New columns: childrenCount_, endReason_ for family sheet
5. Updated language sheet with translations for new fields
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import os

# Paths
INPUT_FILE = '/tmp/gh-issue-solver-1772736320798/ver3/tree.xlsx'
OUTPUT_FILE = '/tmp/gh-issue-solver-1772736320798/ver3/tree.xlsx'
BACKUP_FILE = '/tmp/gh-issue-solver-1772736320798/ver3/tree_backup.xlsx'

# Create backup
shutil.copy(INPUT_FILE, BACKUP_FILE)
print(f"Backup created: {BACKUP_FILE}")

# Load the workbook
wb = load_workbook(INPUT_FILE)

# ============================================================================
# 1. UPDATE PERSON SHEET
# ============================================================================
print("\n=== Updating person sheet ===")

person_sheet = wb['person']
# Get current headers
headers = [cell.value for cell in person_sheet[1]]
print(f"Current headers: {headers}")

# Add new columns if they don't exist
new_person_columns = ['birthPlace_', 'deathPlace_', 'burialPlace_', 'occupation_']
for col_name in new_person_columns:
    if col_name not in headers:
        next_col = len(headers) + 1
        person_sheet.cell(row=1, column=next_col, value=col_name)
        headers.append(col_name)
        print(f"Added column: {col_name}")

# Get column indices
col_indices = {h: i+1 for i, h in enumerate(headers) if h}

# Data for new siblings of Ilya Ulyanov (children of Ульянин Николай Васильевич + Смирнова Анна Алексеевна)
ulyanov_siblings = [
    # Vasily - older brother
    {
        'label_': 'Ульянов Василий Николаевич',
        'sex': 'М',
        'hasFather': 'Ульянин_Николай_Васильевич',
        'hasMother': 'Смирнова_Анна_Алексеевна',
        'birth': '1823',  # approximately 14 years old in 1837
        'birthFull_': '~1823',
        'death': '~1878',
        'deathFull_': '~1878',
        'occupation_': 'приказчик (управляющий соляным делом)'
    },
    # Maria - sister
    {
        'label_': 'Ульянова Мария Николаевна',
        'sex': 'Ж',
        'surName2_': 'Горшкова',
        'hasFather': 'Ульянин_Николай_Васильевич',
        'hasMother': 'Смирнова_Анна_Алексеевна',
        'birth': '~1825',
        'birthFull_': '~1825',
        'death': '?',
        'deathFull_': '?',
        'occupation_': 'домохозяйка'
    },
    # Fedosya - sister
    {
        'label_': 'Ульянова Феодосия Николаевна',
        'sex': 'Ж',
        'hasFather': 'Ульянин_Николай_Васильевич',
        'hasMother': 'Смирнова_Анна_Алексеевна',
        'birth': '~1827',
        'birthFull_': '~1827',
        'death': '1908',
        'deathFull_': '1908',
        'occupation_': ''
    },
]

# Data for siblings of Maria Blank (children of Бланк Александр Дмитриевич + Гроссшопф Анна Ивановна)
blank_siblings = [
    # Dmitry - brother (committed suicide as student)
    {
        'label_': 'Бланк Дмитрий Александрович',
        'sex': 'М',
        'hasFather': 'Бланк_Александр_Дмитриевич',
        'hasMother': 'Гроссшопф_Анна_Ивановна',
        'birth': '1830',
        'birthFull_': '09.09.1830',
        'death': '1850',
        'deathFull_': '19.01.1850',
        'occupation_': 'студент юридического факультета',
        'birthPlace_': 'Российская Империя',
        'deathPlace_': 'Казань'
    },
    # Anna - sister
    {
        'label_': 'Бланк Анна Александровна',
        'sex': 'Ж',
        'surName2_': 'Веретенникова',
        'hasFather': 'Бланк_Александр_Дмитриевич',
        'hasMother': 'Гроссшопф_Анна_Ивановна',
        'birth': '1831',
        'birthFull_': '30.08.1831',
        'death': '?',
        'deathFull_': '?',
        'occupation_': 'домохозяйка'
    },
    # Lyubov - sister
    {
        'label_': 'Бланк Любовь Александровна',
        'sex': 'Ж',
        'surName2_': 'Ардашева/Пономарёва',
        'hasFather': 'Бланк_Александр_Дмитриевич',
        'hasMother': 'Гроссшопф_Анна_Ивановна',
        'birth': '1832',
        'birthFull_': '29.08.1832',
        'death': '?',
        'deathFull_': '?',
        'occupation_': 'домохозяйка'
    },
    # Ekaterina - sister
    {
        'label_': 'Бланк Екатерина Александровна',
        'sex': 'Ж',
        'surName2_': 'Залежская',
        'hasFather': 'Бланк_Александр_Дмитриевич',
        'hasMother': 'Гроссшопф_Анна_Ивановна',
        'birth': '1834',
        'birthFull_': '04.01.1834',
        'death': '?',
        'deathFull_': '?',
        'occupation_': 'домохозяйка'
    },
    # Sophia - sister
    {
        'label_': 'Бланк Софья Александровна',
        'sex': 'Ж',
        'surName2_': 'Лаврова',
        'hasFather': 'Бланк_Александр_Дмитриевич',
        'hasMother': 'Гроссшопф_Анна_Ивановна',
        'birth': '1836',
        'birthFull_': '24.07.1836',
        'death': '?',
        'deathFull_': '?',
        'occupation_': 'домохозяйка'
    },
]

# Update existing persons with new fields
# Ульянин Николай Васильевич
# Find and update the row
for row in range(2, person_sheet.max_row + 1):
    label = person_sheet.cell(row=row, column=col_indices.get('label_', 2)).value
    if label == 'Ульянин Николай Васильевич':
        if 'birthPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['birthPlace_'], value='Нижегородская губерния')
        if 'deathPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['deathPlace_'], value='Астрахань')
        if 'occupation_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['occupation_'], value='портной, мещанин')
        print(f"Updated: {label}")
    elif label == 'Смирнова Анна Алексеевна':
        if 'birthPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['birthPlace_'], value='Астрахань')
        if 'deathPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['deathPlace_'], value='Астрахань')
        if 'occupation_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['occupation_'], value='домохозяйка')
        print(f"Updated: {label}")
    elif label == 'Бланк Александр Дмитриевич':
        if 'birthPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['birthPlace_'], value='Староконстантинов')
        if 'deathPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['deathPlace_'], value='Кокушкино')
        if 'occupation_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['occupation_'], value='врач')
        print(f"Updated: {label}")
    elif label == 'Гроссшопф Анна Ивановна':
        if 'birthPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['birthPlace_'], value='Любек, Германия')
        if 'deathPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['deathPlace_'], value='?')
        if 'occupation_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['occupation_'], value='домохозяйка')
        print(f"Updated: {label}")
    elif label == 'Ульянов Илья Николаевич':
        if 'birthPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['birthPlace_'], value='Астрахань')
        if 'deathPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['deathPlace_'], value='Симбирск')
        if 'occupation_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['occupation_'], value='педагог, инспектор народных училищ')
        print(f"Updated: {label}")
    elif label == 'Бланк Мария Александровна':
        if 'birthPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['birthPlace_'], value='Санкт-Петербург')
        if 'deathPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['deathPlace_'], value='Петроград')
        if 'occupation_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['occupation_'], value='учительница')
        print(f"Updated: {label}")
    elif label == 'Ульянов Владимир Ильич':
        if 'birthPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['birthPlace_'], value='Симбирск')
        if 'deathPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['deathPlace_'], value='Горки')
        if 'burialPlace_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['burialPlace_'], value='Мавзолей, Москва')
        if 'occupation_' in col_indices:
            person_sheet.cell(row=row, column=col_indices['occupation_'], value='революционер, политик')
        print(f"Updated: {label}")

# Add new persons
next_row = person_sheet.max_row + 1

# Refresh column indices after adding new columns
headers = [cell.value for cell in person_sheet[1]]
col_indices = {h: i+1 for i, h in enumerate(headers) if h}

for person_data in ulyanov_siblings + blank_siblings:
    # Generate idA from label
    label = person_data.get('label_', '')
    idA = label.replace(' ', '_')

    # Write idA (leave empty for formula-based calculation in Excel)
    # person_sheet.cell(row=next_row, column=col_indices['idA'], value=None)

    # Write other fields
    for field, value in person_data.items():
        if field in col_indices:
            person_sheet.cell(row=next_row, column=col_indices[field], value=value)

    print(f"Added person: {label}")
    next_row += 1

# ============================================================================
# 2. UPDATE FAMILY SHEET
# ============================================================================
print("\n=== Updating family sheet ===")

family_sheet = wb['family']
family_headers = [cell.value for cell in family_sheet[1]]
print(f"Current family headers: {family_headers}")

# Add new columns if they don't exist
new_family_columns = ['childrenCount_', 'endReason_']
for col_name in new_family_columns:
    if col_name not in family_headers:
        next_col = len(family_headers) + 1
        family_sheet.cell(row=1, column=next_col, value=col_name)
        family_headers.append(col_name)
        print(f"Added column: {col_name}")

family_col_indices = {h: i+1 for i, h in enumerate(family_headers) if h}

# Update existing family entries
for row in range(2, family_sheet.max_row + 1):
    husband = family_sheet.cell(row=row, column=family_col_indices.get('husband', 2)).value
    wife = family_sheet.cell(row=row, column=family_col_indices.get('wife', 3)).value

    if husband == 'Ульянов_Илья_Николаевич' and wife == 'Бланк_Мария_Александровна':
        if 'childrenCount_' in family_col_indices:
            family_sheet.cell(row=row, column=family_col_indices['childrenCount_'], value='8')
        if 'endReason_' in family_col_indices:
            family_sheet.cell(row=row, column=family_col_indices['endReason_'], value='смерть супруга')
        print(f"Updated family: {husband} - {wife}")
    elif husband == 'Бланк_Александр_Дмитриевич' and wife == 'Гроссшопф_Анна_Ивановна':
        if 'childrenCount_' in family_col_indices:
            family_sheet.cell(row=row, column=family_col_indices['childrenCount_'], value='6')
        if 'endReason_' in family_col_indices:
            family_sheet.cell(row=row, column=family_col_indices['endReason_'], value='смерть супруги')
        print(f"Updated family: {husband} - {wife}")
    elif husband == 'Ульянов_Владимир_Ильич' and wife == 'Крупская_Надежда_Константиновна':
        if 'childrenCount_' in family_col_indices:
            family_sheet.cell(row=row, column=family_col_indices['childrenCount_'], value='0')
        if 'endReason_' in family_col_indices:
            family_sheet.cell(row=row, column=family_col_indices['endReason_'], value='смерть супруга')
        print(f"Updated family: {husband} - {wife}")

# Add new family entry for Ульянин + Смирнова
next_family_row = family_sheet.max_row + 1
new_family = {
    'husband': 'Ульянин_Николай_Васильевич',
    'wife': 'Смирнова_Анна_Алексеевна',
    'marriage_': '~1823',
    'locationM_': 'Астрахань',
    'childrenCount_': '5',
    'endReason_': 'смерть супруга'
}

for field, value in new_family.items():
    if field in family_col_indices:
        family_sheet.cell(row=next_family_row, column=family_col_indices[field], value=value)
print(f"Added family: Ульянин_Николай_Васильевич - Смирнова_Анна_Алексеевна")

# ============================================================================
# 3. UPDATE LANGUAGE SHEET
# ============================================================================
print("\n=== Updating language sheet ===")

language_sheet = wb['language']
lang_headers = [cell.value for cell in language_sheet[1]]
print(f"Language headers: {lang_headers}")

lang_col_indices = {h: i+1 for i, h in enumerate(lang_headers) if h}

# New translations to add
new_translations = [
    # Person sheet new fields
    {'type': 'sheet', 'object': 'person', 'name': 'birthPlace_', 'ru': 'место рождения', 'en': 'birthplace'},
    {'type': 'sheet', 'object': 'person', 'name': 'deathPlace_', 'ru': 'место смерти', 'en': 'place of death'},
    {'type': 'sheet', 'object': 'person', 'name': 'burialPlace_', 'ru': 'место захоронения', 'en': 'burial place'},
    {'type': 'sheet', 'object': 'person', 'name': 'occupation_', 'ru': 'профессия', 'en': 'occupation'},
    # Family sheet new fields
    {'type': 'sheet', 'object': 'family', 'name': 'childrenCount_', 'ru': 'число детей', 'en': 'children count'},
    {'type': 'sheet', 'object': 'family', 'name': 'endReason_', 'ru': 'причина окончания', 'en': 'end reason'},
]

# Check which translations already exist
existing_translations = set()
for row in range(2, language_sheet.max_row + 1):
    obj = language_sheet.cell(row=row, column=lang_col_indices.get('object', 2)).value
    name = language_sheet.cell(row=row, column=lang_col_indices.get('name', 3)).value
    if obj and name:
        existing_translations.add(f"{obj}.{name}")

next_lang_row = language_sheet.max_row + 1
for trans in new_translations:
    key = f"{trans['object']}.{trans['name']}"
    if key not in existing_translations:
        for field, value in trans.items():
            if field in lang_col_indices:
                language_sheet.cell(row=next_lang_row, column=lang_col_indices[field], value=value)
        print(f"Added translation: {key}")
        next_lang_row += 1
    else:
        print(f"Translation already exists: {key}")

# Save the workbook
wb.save(OUTPUT_FILE)
print(f"\n=== Saved to {OUTPUT_FILE} ===")
